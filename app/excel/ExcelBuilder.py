import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment
from openpyxl.utils.cell import coordinate_from_string


class ExcelBuilder:
    """
    A class for building and modifying Excel files using the openpyxl library.
    """

    def __init__(self, file_name):
        """
        Initialize the ExcelBuilder with a specified file name.

        Parameters:
            file_name (str): The name of the Excel file.
        """
        self.file_name = file_name
        self.workbook = openpyxl.load_workbook(f'{self.file_name}.xlsx')

    def add_column_to_excel(self, values):
        """
        Add a column to the 'Dados' sheet with the specified values.

        Parameters:
            values (list): The list of values to be added to the column.
        """
        worksheet = self.workbook['Dados']

        for i, valor in enumerate(values):
            worksheet.cell(row=i + 2, column=16, value=values[i])

        self.save()

    def add_countif_formula(self, target_column, target_value, result_cell):
        """
        Add a COUNTIF formula to the specified cell in the 'Dados' sheet.

        Parameters:
            target_column (int): The column index for the COUNTIF range.
            target_value (str): The value to count in the specified range.
            result_cell (str): The target cell for the COUNTIF result.
        """
        worksheet = self.workbook['Dados']
        formula = f'=COUNTIF({get_column_letter(target_column)}:{get_column_letter(target_column)},{target_value})'

        cell = worksheet[result_cell]
        cell.value = formula
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '0'

    def add_if_formula(self, sheet_name, target_column, target_row, condition_column, condition_row,
                       if_success, if_fail, result_cell):
        """
        Add an IF formula to the specified cell in the specified sheet.

        Parameters:
            sheet_name (str): The name of the sheet where the formula should be added.
            target_column (int): The column index for the target cell.
            target_row (int): The row index for the target cell.
            condition_column (int): The column index for the condition cell.
            condition_row (int): The row index for the condition cell.
            if_success (str): The value to be placed in the result cell if the condition is true.
            if_fail (str): The value to be placed in the result cell if the condition is false.
            result_cell (str): The target cell for the IF formula result.
        """
        worksheet = self.workbook[sheet_name]
        formula = (f'=IF({get_column_letter(target_column)}{target_row}='
                   f'{get_column_letter(condition_column)}{condition_row},"{if_success}","{if_fail}")')

        cell = worksheet[result_cell]
        cell.value = formula
        cell.alignment = Alignment(horizontal='center', vertical='center')

        self.save()

    def add_title_header(self, sheet_name, header_titles):
        """
        Add title headers to the specified sheet.

        Parameters:
            sheet_name (str): The name of the sheet where the title headers should be added.
            header_titles (dict): A dictionary mapping cell coordinates to header texts.
        """
        worksheet = self.workbook[sheet_name]

        for cell_coord, header_text in header_titles.items():
            col, row = coordinate_from_string(cell_coord)

            cell = worksheet.cell(row=row, column=column_index_from_string(col))
            cell.value = header_text

        self.save()

    def save(self):
        """Save the workbook to the file."""
        self.workbook.save(f'{self.file_name}.xlsx')

    def close(self):
        """Close the workbook."""
        self.workbook.close()
