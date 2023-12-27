import openpyxl
import pyperclip as pc
import pandas as pd
import io


class ExcelProcessor:
    """
    A class for processing Excel data, including reading from clipboard, writing to Excel files, and copying data between
    Excel files and sheets.
    """

    def __init__(self, file_name, filename_workbook=False):
        """
        Initialize the ExcelProcessor.

        Parameters:
            file_name (str): The name of the Excel file.
            filename_workbook (bool): Whether to load the workbook upon initialization. Default is False.
        """
        self.file_name = file_name
        if filename_workbook:
            self.workbook = openpyxl.load_workbook(f'{self.file_name}.xlsx')

    @staticmethod
    def read_clipboard_to_dataframe():
        """
        Read data from the clipboard and convert it to a Pandas DataFrame.

        Returns:
            pd.DataFrame: The DataFrame created from the clipboard data.
        """
        data = pc.paste()
        return pd.read_csv(io.StringIO(data), sep='\t')

    def write_dataframe_to_excel(self, dataframe, sheet_name='Dados'):
        """
        Write a Pandas DataFrame to an Excel file.

        Parameters:
            dataframe (pd.DataFrame): The DataFrame to be written.
            sheet_name (str): The name of the sheet to write the DataFrame to. Default is 'Dados'.
        """
        writer = pd.ExcelWriter(f'{self.file_name}.xlsx', engine='openpyxl')
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        writer._save()
        self.workbook = openpyxl.load_workbook(f'{self.file_name}.xlsx')

    @staticmethod
    def copy_excel_data(source_file, source_sheet, target_file, target_sheet, start_row, end_row, start_col, end_col):
        """
        Copy a range of data from one Excel file/sheet to another.

        Parameters:
            source_file (str): The source Excel file.
            source_sheet (str): The source sheet in the source Excel file.
            target_file (str): The target Excel file.
            target_sheet (str): The target sheet in the target Excel file.
            start_row (int): The starting row index for the data to copy.
            end_row (int): The ending row index for the data to copy.
            start_col (str): The starting column letter for the data to copy.
            end_col (str): The ending column letter for the data to copy.
        """
        df = pd.read_excel(source_file, sheet_name=source_sheet)
        selection = df.loc[start_row:end_row, start_col:end_col]

        new_spreadsheet = pd.DataFrame(selection)

        writer = pd.ExcelWriter(target_file, engine='openpyxl')
        new_spreadsheet.to_excel(writer, sheet_name=target_sheet, index=False)
        writer._save()

    def get_max_row(self, sheet_name):
        """
        Get the maximum row number in a specified sheet.

        Parameters:
            sheet_name (str): The name of the sheet.

        Returns:
            int: The maximum row number.
        """
        worksheet = self.workbook[sheet_name]
        return worksheet.max_row

    def save(self):
        """Save the workbook to the file."""
        self.workbook.save(f'{self.file_name}.xlsx')

    def close(self):
        """Close the workbook."""
        self.workbook.close()
