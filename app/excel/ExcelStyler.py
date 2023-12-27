import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelStyler:
    """
    A class for styling Excel sheets using the openpyxl library.
    """

    def __init__(self, file_name):
        """
        Initialize the ExcelStyler with a specified file name.

        Parameters:
            file_name (str): The name of the Excel file.
        """
        self.file_name = file_name
        self.workbook = openpyxl.load_workbook(f'{file_name}.xlsx')

    def apply_style_to_header(self, sheet_name, column_number):
        """
        Apply a predefined style to the header cell of a specified column.

        Parameters:
            sheet_name (str): The name of the sheet.
            column_number (int): The number of the column to style.
        """
        header_cell = self.workbook[sheet_name][f"{get_column_letter(column_number)}1"]
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.fill = PatternFill(start_color='FFA600', end_color='FFA600', fill_type='solid')

        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        header_cell.border = border

        self.save()

    def set_column_width(self, sheet_name, column_number, width=""):
        """
        Set the width of a specified column in a sheet.

        Parameters:
            sheet_name (str): The name of the sheet.
            column_number (int): The number of the column to set the width.
            width (int): The width to set for the column (optional). If not provided, it is determined based on content.
        """
        if width == "":
            width = max(len(str(cell.value)) + 4
                        for cell in self.workbook[sheet_name][get_column_letter(column_number)])

        self.workbook[sheet_name].column_dimensions[get_column_letter(column_number)].width = width
        self.save()

    def set_column_alignment(self, sheet_name, column_number, horizontal='center', vertical='center'):
        """
        Set the alignment of cells in a specified column in a sheet.

        Parameters:
            sheet_name (str): The name of the sheet.
            column_number (int): The number of the column to set alignment.
            horizontal (str): Horizontal alignment ('center' by default).
            vertical (str): Vertical alignment ('center' by default).
        """
        worksheet = self.workbook[sheet_name]

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                       min_col=column_number, max_col=column_number):
            for cell in row:
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

        self.save()

    def apply_number_format(self, sheet_name, column_number, row_number, number_format="General"):
        """
        Apply a numeric format to a specific cell.

        Parameters:
            sheet_name (str): The name of the sheet.
            column_number (int): The number of the column.
            row_number (int): The number of the row.
            number_format (str): The desired numeric format ("General" by default).
        """
        cell = self.workbook[sheet_name][f"{get_column_letter(column_number)}{row_number}"]
        cell.number_format = number_format
        self.save()

    def save(self):
        """Save the workbook to the file."""
        self.workbook.save(f'{self.file_name}.xlsx')

    def close(self):
        """Close the workbook."""
        self.workbook.close()
