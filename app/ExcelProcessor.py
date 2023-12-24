import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import pyperclip as pc
import pandas as pd
import io


class ExcelProcessor:
    def __init__(self, file_name, filename_workbook=False):
        self.file_name = file_name
        if filename_workbook:
            self.workbook = openpyxl.load_workbook(f'{self.file_name}.xlsx')

    @staticmethod
    def read_clipboard_to_dataframe():
        data = pc.paste()
        return pd.read_csv(io.StringIO(data), sep='\t')

    def write_dataframe_to_excel(self, dataframe, sheet_name='Dados'):
        writer = pd.ExcelWriter(f'{self.file_name}.xlsx', engine='openpyxl')
        dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        writer._save()
        self.workbook = openpyxl.load_workbook(f'{self.file_name}.xlsx')

    @staticmethod
    def copy_excel_data(source_file, source_sheet, target_file, target_sheet, start_row, end_row, start_col, end_col):
        df = pd.read_excel(source_file, sheet_name=source_sheet)
        selection = df.loc[start_row:end_row, start_col:end_col]

        new_spreadsheet = pd.DataFrame(selection)

        writer = pd.ExcelWriter(target_file, engine='openpyxl')
        new_spreadsheet.to_excel(writer, sheet_name=target_sheet, index=False)
        writer._save()

    def add_column_to_excel(self, values):
        worksheet = self.workbook['Dados']

        for i, valor in enumerate(values):
            worksheet.cell(row=i + 2, column=16, value=values[i])

        self.save()

    def add_countif_formula(self, target_column, target_value, result_cell):
        worksheet = self.workbook['Dados']
        formula = f'=COUNTIF({get_column_letter(target_column)}:{get_column_letter(target_column)},{target_value})'

        cell = worksheet[result_cell]
        cell.value = formula
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.number_format = '0'

    def add_if_formula(self, sheet_name, target_column, target_row, condition_column, condition_row,
                       if_sucess, if_fail, result_cell):
        worksheet = self.workbook[sheet_name]
        formula = (f'=IF({get_column_letter(target_column)}{target_row}='
                   f'{get_column_letter(condition_column)}{condition_row},"{if_sucess}","{if_fail}")')

        cell = worksheet[result_cell]
        cell.value = formula
        cell.alignment = Alignment(horizontal='center', vertical='center')

        self.save()

    def add_title_header(self, sheet_name, header_titles):
        worksheet = self.workbook[sheet_name]

        for cell_coord, header_text in header_titles.items():
            col, row = coordinate_from_string(cell_coord)

            cell = worksheet.cell(row=row, column=column_index_from_string(col))
            cell.value = header_text

        self.save()

    def get_max_row(self, sheet_name):
        worksheet = self.workbook[sheet_name]
        return worksheet.max_row

    def save(self):
        self.workbook.save(f'{self.file_name}.xlsx')

    def close(self):
        self.workbook.close()
