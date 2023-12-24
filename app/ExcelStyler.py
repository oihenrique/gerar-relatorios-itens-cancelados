import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelStyler:
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = openpyxl.load_workbook(f'{file_name}.xlsx')

    def apply_style_to_header(self, sheet_name, column_number):
        header_cell = self.workbook[sheet_name][f"{get_column_letter(column_number)}1"]
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.fill = PatternFill(start_color='FFA600', end_color='FFA600', fill_type='solid')

        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))
        header_cell.border = border

        self.save()

    def set_column_width(self, sheet_name, column_number, width=""):
        if width == "":
            width = max(len(str(cell.value)) + 4
                        for cell in self.workbook[sheet_name][get_column_letter(column_number)])

        self.workbook[sheet_name].column_dimensions[get_column_letter(column_number)].width = width
        self.save()

    def set_column_alignment(self, sheet_name, column_number, horizontal='center', vertical='center'):
        worksheet = self.workbook[sheet_name]

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                       min_col=column_number, max_col=column_number):
            for cell in row:
                cell.alignment = Alignment(horizontal=horizontal, vertical=vertical)

        self.save()

    def apply_number_format(self, sheet_name, column_number, row_number, number_format="General"):
        """
        Aplica um formato numérico a uma célula específica.

        :param sheet_name: Nome da planilha.
        :param column_number: Número da coluna.
        :param row_number: Número da linha.
        :param number_format: Formato numérico desejado (por padrão, "General").
        """
        cell = self.workbook[sheet_name][f"{get_column_letter(column_number)}{row_number}"]
        cell.number_format = number_format
        self.save()

    def save(self):
        self.workbook.save(f'{self.file_name}.xlsx')

    def close(self):
        self.workbook.close()
